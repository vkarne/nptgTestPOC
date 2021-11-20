# This utility file used to read, write data to a exel file

import openpyxl


def get_row_count(file, sheet_name):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet_name]
    return ws.max_row


def get_col_count(file, sheet_name):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet_name]
    return ws.max_column


def read_data(file, sheet_name, row_num, column_num):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet_name]
    cd = ws.cell(row=row_num, column=column_num)
    return cd.value


def write_data(file, sheet_name, row_num, column_num, data):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet_name]
    ws.cell(row=row_num, column=column_num).value = data
    wb.save(file)

